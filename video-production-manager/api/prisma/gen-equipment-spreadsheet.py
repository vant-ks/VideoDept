#!/usr/bin/env python3
"""Generate an equipment I/O spec spreadsheet from equipment-data.json.

Columns: ID | Manufacturer | Model | Category | (Port N Type | Port N Label) x 20
Port type cells have data validation dropdowns from the connector type list.
Inputs come first (labeled with arrow ← or the label itself), then outputs (→ or label).
"""
import json, os, sys
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── paths ────────────────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_FILE  = os.path.join(SCRIPT_DIR, "equipment-data.json")
OUT_FILE   = os.path.join(SCRIPT_DIR, "equipment-io-spec.xlsx")

MAX_PORTS  = 20   # columns per direction group → 20 total port slots

CONNECTOR_TYPES = [
    "HDMI 1.4", "HDMI 2.0", "HDMI 2.1",
    "3G-SDI", "6G-SDI", "12G-SDI", "BNC REF",
    "DP 1.1", "DP 1.2", "DP 1.4",
    "NDI", "USB-C", "NETWORK (RJ45)",
    "OPTICON DUO", "OPTICON QUAD",
    "SMPTE FIBER",
    "LC - FIBER (SM)", "ST - FIBER (SM)", "SC - FIBER (SM)",
    "LC - FIBER (MM)", "ST - FIBER (MM)", "SC - FIBER (MM)",
    "XLR", "DMX",
    # legacy / other values already in data
    "HDMI", "SDI", "DisplayPort", "Ethernet", "Network (RJ45)",
    "3.5mm", "Headphone", "Mini XLR", "Reference", "Reference (BNC)",
    "TRS", "RCA", "USB", "SFP Optical", "SMPTE Fiber",
    "IP", "SD Card", "Power",
]

# ── colour palette ───────────────────────────────────────────────────────────
HDR_BG  = "1E2A38"   # dark navy  (header row)
HDR_FG  = "FFFFFF"
GRP_IN  = "0D3B66"   # deep blue  (Inputs group header)
GRP_OUT = "6B3A26"   # burnt-sienna (Outputs group header)
TYPE_IN  = "DDE8F7"  # light blue  (type cells – inputs)
TYPE_OUT = "FAE8DC"  # light peach (type cells – outputs)
LBL_IN   = "EFF5FC"
LBL_OUT  = "FDF4EF"
META_BG  = "2C3A4A"
ROW_ALT  = "F5F8FB"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def bold(color="000000", sz=10):
    return Font(bold=True, color=color, size=sz)

def thin_border():
    s = Side(style="thin", color="C0C0C0")
    return Border(left=s, right=s, top=s, bottom=s)

# ── load data ────────────────────────────────────────────────────────────────
with open(DATA_FILE) as f:
    equipment = json.load(f)

print(f"Loaded {len(equipment)} equipment entries.")

# ── build workbook ───────────────────────────────────────────────────────────
wb = Workbook()

# ── Sheet 2: connector type reference list ───────────────────────────────────
ref_ws = wb.active
ref_ws.title = "_ConnectorTypes"
ref_ws["A1"] = "Connector Types"
ref_ws["A1"].font = bold(sz=11)
for i, ct in enumerate(CONNECTOR_TYPES, start=2):
    ref_ws.cell(row=i, column=1, value=ct)
ref_ws.column_dimensions["A"].width = 28

# ── Sheet 1: Equipment I/O ───────────────────────────────────────────────────
ws = wb.create_sheet("Equipment IO", 0)

# ── column layout ────────────────────────────────────────────────────────────
# cols 1-4: meta   |  5-(4+2*MAX_PORTS): Port 1 Type, Port 1 Label, Port 2 Type …
META_COLS = 4
PORT_START = META_COLS + 1   # column index of first port column (1-based)

total_cols = META_COLS + MAX_PORTS * 2

# ── row 1: top-level group labels ────────────────────────────────────────────
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=META_COLS)
ws.cell(1, 1, "EQUIPMENT").fill = fill(META_BG)
ws.cell(1, 1).font = bold("FFFFFF", 11)
ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")

for i in range(MAX_PORTS):
    type_col = PORT_START + i * 2
    ws.cell(1, type_col, f"PORT {i+1}").fill = fill(HDR_BG)
    ws.cell(1, type_col).font = bold("FFFFFF", 9)
    ws.cell(1, type_col).alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=1, start_column=type_col, end_row=1, end_column=type_col + 1)

# ── row 2: column headers ────────────────────────────────────────────────────
meta_headers = ["ID", "Manufacturer", "Model", "Category"]
for c, h in enumerate(meta_headers, start=1):
    cell = ws.cell(2, c, h)
    cell.fill = fill(HDR_BG)
    cell.font = bold("FFFFFF", 10)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border()

for i in range(MAX_PORTS):
    type_col  = PORT_START + i * 2
    label_col = type_col + 1
    tc = ws.cell(2, type_col, "Type")
    tc.fill = fill(GRP_IN if i < 10 else GRP_OUT)
    tc.font = bold("FFFFFF", 9)
    tc.alignment = Alignment(horizontal="center")
    tc.border = thin_border()
    lc = ws.cell(2, label_col, "Label")
    lc.fill = fill(GRP_IN if i < 10 else GRP_OUT)
    lc.font = bold("FFFFFF", 9)
    lc.alignment = Alignment(horizontal="center")
    lc.border = thin_border()

ws.row_dimensions[1].height = 18
ws.row_dimensions[2].height = 18
ws.freeze_panes = "A3"

# ── data validation ──────────────────────────────────────────────────────────
# Reference the _ConnectorTypes sheet for the dropdown source
ct_ref_end = len(CONNECTOR_TYPES) + 1   # row 2 to row N in _ConnectorTypes
dv_formula = f"_ConnectorTypes!$A$2:$A${ct_ref_end}"

# We'll add one DataValidation per Type column (openpyxl can combine ranges)
# Build a single DV that covers all type columns across all data rows
DATA_ROWS_START = 3
DATA_ROWS_END   = DATA_ROWS_START + len(equipment) - 1

dv = DataValidation(
    type="list",
    formula1=f"={dv_formula}",
    allow_blank=True,
    showDropDown=False,   # False = show the dropdown arrow
    showErrorMessage=True,
    error="Choose from the list or leave blank.",
    errorTitle="Invalid Connector Type",
)
ws.add_data_validation(dv)

# ── populate equipment rows ───────────────────────────────────────────────────
for row_idx, item in enumerate(equipment, start=DATA_ROWS_START):
    alt = (row_idx % 2 == 0)
    bg  = ROW_ALT if alt else "FFFFFF"

    # meta columns
    for c, val in enumerate([
        item.get("id",""),
        item.get("manufacturer",""),
        item.get("model",""),
        item.get("category",""),
    ], start=1):
        cell = ws.cell(row_idx, c, val)
        cell.fill   = fill(bg)
        cell.border = thin_border()
        cell.alignment = Alignment(vertical="center")

    # combine inputs then outputs into a flat port list (max MAX_PORTS slots)
    ports = []
    for p in item.get("inputs", []):
        ports.append((p.get("type",""), p.get("label","") + " ←"))
    for p in item.get("outputs", []):
        ports.append((p.get("type",""), p.get("label","") + " →"))

    for slot in range(MAX_PORTS):
        type_col  = PORT_START + slot * 2
        label_col = type_col + 1
        is_input_zone = slot < 10

        if slot < len(ports):
            ptype, plabel = ports[slot]
        else:
            ptype, plabel = "", ""

        tbg = (TYPE_IN if is_input_zone else TYPE_OUT) if ptype else bg
        lbg = (LBL_IN  if is_input_zone else LBL_OUT)  if ptype else bg

        tc = ws.cell(row_idx, type_col,  ptype)
        tc.fill   = fill(tbg)
        tc.border = thin_border()
        tc.alignment = Alignment(horizontal="center", vertical="center")
        dv.add(tc)   # attach dropdown

        lc = ws.cell(row_idx, label_col, plabel)
        lc.fill   = fill(lbg)
        lc.border = thin_border()
        lc.alignment = Alignment(vertical="center")

# ── column widths ─────────────────────────────────────────────────────────────
ws.column_dimensions[get_column_letter(1)].width = 38   # ID
ws.column_dimensions[get_column_letter(2)].width = 16   # Manufacturer
ws.column_dimensions[get_column_letter(3)].width = 28   # Model
ws.column_dimensions[get_column_letter(4)].width = 16   # Category

for i in range(MAX_PORTS):
    type_col  = PORT_START + i * 2
    label_col = type_col + 1
    ws.column_dimensions[get_column_letter(type_col)].width  = 16
    ws.column_dimensions[get_column_letter(label_col)].width = 26

# ── row heights ───────────────────────────────────────────────────────────────
for r in range(DATA_ROWS_START, DATA_ROWS_END + 1):
    ws.row_dimensions[r].height = 16

# ── save ──────────────────────────────────────────────────────────────────────
wb.save(OUT_FILE)
print(f"✅  Spreadsheet saved to: {OUT_FILE}")
print(f"    {len(equipment)} equipment rows | {MAX_PORTS} port slots each")
